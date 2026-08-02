# make_report.py
# 出库入库明细报表：按日期范围生成每个花型的每日出库/入库明细。
import os
from datetime import datetime, timedelta
from sqlalchemy import text
from mysql_conn import engine
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"D:\店铺\出库入库明细报告"

NUM_FMT = "0.0"
HEADER_FILL = "D9E1F2"


def _ensure_output_dir():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)


def _to_date(value):
    if value is None:
        raise ValueError("日期不能为空")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return datetime.strptime(value, '%Y%m%d').date()
    if hasattr(value, 'date'):
        return value.date()
    return value


def generate_inout_detail_report(start_date, end_date):
    """
    生成指定日期范围的出库入库明细报表。

    参数:
        start_date: 开始日期 (date 对象或 'YYYY-MM-DD' 字符串)
        end_date:  结束日期 (date 对象或 'YYYY-MM-DD' 字符串)

    返回:
        生成的 Excel 文件路径

    报表结构（每个花型占 2 行）：
        第1行  标题（合并整行）：{开始日期YYMMDD}-{结束日期YYMMDD}出库入库明细
        第2行  B2='日期'，D列起为日期数字（1..N）
        第3行  A='{M}月{D}日库存' B='花型名称' 倒数第2列='出库总数' 末列='库存余额'
        数据行  出库行(A期初/B花型/C'出库米数'/每日出库/出库总数/库存余额)，
                入库行(A/B空/C'入库米数'/每日入库/入库总数/空)
    入库按生效日期（effect_date）归集，与库存重算口径一致。
    """
    start_dt = _to_date(start_date)
    end_dt = _to_date(end_date)
    if start_dt > end_dt:
        raise ValueError("开始日期不能晚于结束日期")

    days = []
    d = start_dt
    while d <= end_dt:
        days.append(d)
        d += timedelta(days=1)

    _ensure_output_dir()

    with engine.connect() as conn:
        # 花型列表（未删除，按名称排序）
        flowers = [r[0] for r in conn.execute(text(
            "SELECT flower FROM product_cost WHERE is_deleted = 0 ORDER BY flower"
        )).fetchall()]
        if not flowers:
            raise ValueError("请先添加花型")

        # 期初库存 = 开始日期前一天的快照
        prev_date = start_dt - timedelta(days=1)
        prev_label = f"{prev_date.month}月{prev_date.day}日库存"
        init_map = {
            r[0]: float(r[1]) for r in conn.execute(text(
                "SELECT flower, stock FROM inventory_snapshot WHERE snapshot_date = :d"
            ), {"d": prev_date}).fetchall()
        }

        # 每日出库（日报缓存）
        out_map = {
            (r[0], r[1]): float(r[2]) for r in conn.execute(text(
                "SELECT report_date, flower, total_meters FROM daily_report_cache "
                "WHERE report_date BETWEEN :s AND :e"
            ), {"s": start_dt, "e": end_dt}).fetchall()
        }

        # 每日入库（按生效日期归集，历史补录不落到执行当天）
        in_map = {
            (r[0], r[1]): float(r[2]) for r in conn.execute(text(
                "SELECT COALESCE(effect_date, DATE(created_at)) AS d, flower, SUM(change_qty) "
                "FROM inventory_log "
                "WHERE change_type = '入库' "
                "  AND COALESCE(effect_date, DATE(created_at)) BETWEEN :s AND :e "
                "GROUP BY d, flower"
            ), {"s": start_dt, "e": end_dt}).fetchall()
        }

    blocks = []
    total_out_all = 0.0
    total_in_all = 0.0
    for flower in flowers:
        daily_out = [out_map.get((day, flower), 0.0) for day in days]
        daily_in = [in_map.get((day, flower), 0.0) for day in days]
        out_total = sum(daily_out)
        in_total = sum(daily_in)
        init = init_map.get(flower, 0.0)
        blocks.append({
            'flower': flower, 'init': init,
            'daily_out': daily_out, 'daily_in': daily_in,
            'out_total': out_total, 'in_total': in_total,
            'balance': init + in_total - out_total,
        })
        total_out_all += out_total
        total_in_all += in_total

    if total_out_all + total_in_all == 0:
        raise ValueError("所选日期范围内无数据")

    # ============================================================
    # 写 Excel
    # ============================================================
    n = len(days)
    ncols = n + 5  # A期初 + B花型 + C行标 + N天 + 出库总数 + 库存余额
    last_col = get_column_letter(ncols)

    wb = Workbook()
    ws = wb.active
    ws.title = "出库入库明细"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # 第1行：标题（合并整行）
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"].value = f"{start_dt.strftime('%y%m%d')}-{end_dt.strftime('%y%m%d')}出库入库明细"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # 第2行：日期（B2='日期'，D列起为日期数字）
    ws["B2"] = "日期"
    ws["B2"].alignment = center
    for i, day in enumerate(days):
        ws.cell(row=2, column=4 + i, value=day.day).alignment = center

    # 第3行：表头
    ws.cell(row=3, column=1, value=prev_label)
    ws.cell(row=3, column=2, value="花型名称")
    ws.cell(row=3, column=ncols - 1, value="总出入库")
    ws.cell(row=3, column=ncols, value="库存余额")
    for col in (1, 2, ncols - 1, ncols):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
    red_font = Font(color="FF0000")
    # 数据行：每个花型 2 行（出库行 / 入库行）
    row = 4
    for blk in blocks:
        row_out, row_in = row, row + 1

        # A 期初库存（合并 2 行）
        ws.merge_cells(start_row=row_out, start_column=1, end_row=row_in, end_column=1)
        a_cell = ws.cell(row=row_out, column=1, value=blk['init'])
        a_cell.number_format = NUM_FMT
        a_cell.alignment = right

        # B 花型名称（合并 2 行）
        ws.merge_cells(start_row=row_out, start_column=2, end_row=row_in, end_column=2)
        ws.cell(row=row_out, column=2, value=blk['flower']).alignment = center

        # C 行标
        ws.cell(row=row_out, column=3, value="出库米数").alignment = center
        ws.cell(row=row_in, column=3, value="入库米数").alignment = center
        ws.cell(row=row_out, column=3).font = red_font

        # 每日出库 / 入库
        for i, (out_v, in_v) in enumerate(zip(blk['daily_out'], blk['daily_in'])):
            o_cell = ws.cell(row=row_out, column=4 + i, value=out_v)
            o_cell.number_format = NUM_FMT
            o_cell.alignment = right
            i_cell = ws.cell(row=row_in, column=4 + i, value=in_v)
            i_cell.number_format = NUM_FMT
            i_cell.alignment = right
            o_cell.font = red_font

            # 出库总数 / 入库总数（倒数第2列，加粗）
        out_t = ws.cell(row=row_out, column=ncols - 1, value=blk['out_total'])
        out_t.number_format = NUM_FMT
        out_t.font = bold_font
        out_t.alignment = right
        in_t = ws.cell(row=row_in, column=ncols - 1, value=blk['in_total'])
        in_t.number_format = NUM_FMT
        in_t.font = bold_font
        in_t.alignment = right
        out_t.font = red_font

        # 库存余额（末列，合并 2 行，加粗）
        ws.merge_cells(start_row=row_out, start_column=ncols, end_row=row_in, end_column=ncols)
        bal_cell = ws.cell(row=row_out, column=ncols, value=blk['balance'])
        bal_cell.number_format = NUM_FMT
        bal_cell.font = bold_font
        bal_cell.alignment = right

        row += 2

    # 固定列宽：紧凑布局，30 天数据在常见屏幕下无需横向滚动
    ws.column_dimensions['A'].width = 11  # 期初库存
    ws.column_dimensions['B'].width = 12  # 花型名称
    ws.column_dimensions['C'].width = 8   # 出库/入库标签
    for i in range(n):
        ws.column_dimensions[get_column_letter(4 + i)].width = 6  # 每日出库/入库
    ws.column_dimensions[get_column_letter(ncols - 1)].width = 8   # 出库总数
    ws.column_dimensions[get_column_letter(ncols)].width = 10      # 库存余额

    filename = f"{start_dt.strftime('%y%m%d')}-{end_dt.strftime('%y%m%d')}出库入库明细.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    s = sys.argv[1] if len(sys.argv) > 1 else '2026-07-01'
    e = sys.argv[2] if len(sys.argv) > 2 else '2026-07-31'
    print(generate_inout_detail_report(s, e))
