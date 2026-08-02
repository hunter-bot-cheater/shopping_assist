# make_daily.py
import pandas as pd
import os
import re
from datetime import datetime, timedelta
from sqlalchemy import text
from mysql_conn import engine
from import_order import orders, PLATFORM_NAMES, PLATFORM_DOUYIN, extract_flower_from_spec
from inventory_service import deduct_stock, get_missing_report_dates

# ============================
# 配置
# ============================
OUTPUT_DIR = r"D:\店铺\日报"
POSTAGE_PER_ORDER = 2.5

# 汇总表指标列（顺序固定：订单数、成本、米数、营业额、快递费、盈利）
SUMMARY_METRICS = ['订单数', '成本', '米数', '营业额', '快递费', '盈利']
# 花型汇总左右分列展示的平台（新增平台在此追加）
SUMMARY_PLATFORMS = ('拼多多', '淘宝', '抖音')
# 平台汇总表始终展示的平台（无数据补 0）
PLATFORM_SUMMARY_NAMES = ('拼多多', '淘宝', '抖音')
# 已取消/已关闭的订单状态（拼多多=已取消，抖音=已关闭；淘宝=交易关闭单独处理）
CANCELLED_STATUSES = ('已取消', '已关闭')


# ============================
# 工具函数
# ============================
def ensure_output_dir():
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)


def extract_meter_from_spec(spec):
    """提取购买米数（只识别逗号前的长度，避免把宽幅识别进去）"""
    if pd.isna(spec):
        return 0
    spec = str(spec).strip()
    # 全角字符转半角（处理中文输入法产生的全角标点和数字）
    full_to_half = str.maketrans({
        '，': ',', '．': '.', '（': '(', '）': ')',
        '０': '0', '１': '1', '２': '2', '３': '3', '４': '4',
        '５': '5', '６': '6', '７': '7', '８': '8', '９': '9',
    })
    spec = spec.translate(full_to_half)

    # 只取逗号前内容
    first_part = spec.split(',')[1] if ',' in spec else spec
    first_part = re.split(r'[（(]', first_part)[0]
    first_part = first_part.strip()

    # 中文数字
    chinese_map = {
        "半米": 0.5, "半": 0.5,
        "一米": 1, "二米": 2, "两米": 2, "三米": 3, "四米": 4, "五米": 5,
    }
    for k, v in chinese_map.items():
        if k in first_part:
            return v

    # 数字米
    match = re.search(r'(\d+\.?\d*)\s*(米|m)', first_part, re.I)
    if match:
        return float(match.group(1))

    # 抖音规格兜底：米数常在分号后（如「花型;二米（说明）」），
    # 且可能被前面的（门幅）括号截断；分号不出现在拼多多/淘宝规格中，不影响既有结果
    for seg in re.split(r'[;；]', spec):
        seg2 = re.split(r'[（(]', seg)[0].strip()
        for k, v in chinese_map.items():
            if k in seg2:
                return v
        match = re.search(r'(\d+\.?\d*)\s*(米|m)', seg2, re.I)
        if match:
            return float(match.group(1))
    return 0


def load_cost_map(report_date=None):
    """
    按报表日期加载成本
    规则：未删除 或 删除生效日期 > 报表日期（即删除日期之后才隐藏）
    """
    try:
        if not report_date:
            report_date = datetime.now().strftime("%Y-%m-%d")

        sql = """
              SELECT flower, cost_per_meter
              FROM product_cost
              WHERE is_deleted = 0
                 OR (is_deleted = 1 AND delete_effect_date > %s)
              """
        df = pd.read_sql(sql, engine, params=(report_date,))
        return dict(zip(df['flower'], df['cost_per_meter']))
    except Exception as e:
        print(f"加载成本表失败: {e}")
        return {}
# 花型标准名映射：将同花型的不同叫法统一到成本表名称
# （含淘宝商品名与成本表差异、抖音规格别名），日报/月报共用
FLOWER_ALIAS_MAP = {
    '白底乱纹': '白底乱纹',
    '白底黑色条纹': '白底乱纹',
    "几何乱纹": "白底乱纹",
    # 淘宝商品名 → 成本表花型
    '蓝色唐人': '唐人蓝色',
    '紫色唐人': '唐人紫色',
    '皮粉唐人': '唐人粉色',
    '粉色唐人': '唐人粉色',
    '黄色唐人': '黄色小唐人',
    '水墨风蝴蝶': '水墨蝴蝶',
    '雅韵青花': '青花雅韵',
    '三色拼读': '三色拼写',
    "黑底条纹": "城市条纹",
    "黑白玄纹": "玄纹密码",
    "条纹哪吒": "黄底哪吒",
    "腰果花": "灰白底腰果花",
    # 抖音规格别名 → 成本表花型
    '绿野浮蝶': '绿影浮蝶',
    '紫底小唐人': '唐人紫色',
    '白底条纹': '白底乱纹',
    '蓝底小唐人': '唐人蓝色',
    '粉底小唐人': '唐人粉色',
}

# ============================
# 花型匹配与计算（日报/区间报告共用，保证口径一致）
# ============================
def assign_flowers(df, cost_flowers):
    """四层花型匹配：规格提取 → 商品名 contains → 商品+规格拼接 contains → 关键词规则。

    所有行都会得到花型，无法匹配的填 '未匹配'（不过滤）。
    输入 df 需含 'product' 与 'product_spec' 列。
    日报/区间报告/退款明细共用，保证花型口径一致。
    """
    df = df.copy()
    cost_flowers = set(cost_flowers)

    # 第一步：从规格提取（先做别名归一，再做成本表匹配）
    df['spec_flower'] = df['product_spec'].apply(extract_flower_from_spec)
    df['spec_flower'] = df['spec_flower'].replace(FLOWER_ALIAS_MAP)
    df['花型'] = df['spec_flower'].apply(lambda x: x if x in cost_flowers else None)

    # 第二步：从商品名称匹配
    unmatched_mask = df['花型'].isna()
    if unmatched_mask.any():
        for flower in cost_flowers:
            still_unmatched = df['花型'].isna()
            if not still_unmatched.any():
                break
            contains = df.loc[still_unmatched, 'product'].str.contains(flower, na=False, case=False)
            if contains.any():
                df.loc[still_unmatched & contains, '花型'] = flower

    # 第三步：商品+规格拼接匹配
    unmatched_mask = df['花型'].isna()
    if unmatched_mask.any():
        df['concat_field'] = df['product'].fillna('') + '-' + df['product_spec'].fillna('')
        for flower in cost_flowers:
            still_unmatched = df['花型'].isna()
            if not still_unmatched.any():
                break
            contains = df.loc[still_unmatched, 'concat_field'].str.contains(flower, na=False, case=False)
            if contains.any():
                df.loc[still_unmatched & contains, '花型'] = flower
        df.drop(columns=['concat_field'], inplace=True)

    # 第四步：关键词映射
    unmatched_mask = df['花型'].isna()
    if unmatched_mask.any():
        mapping_rules = [
            {'product_keyword': '唐人', 'spec_color_keyword': '紫色', 'target_flower': '唐人紫色'},
            {'product_keyword': '唐人', 'spec_color_keyword': '蓝色', 'target_flower': '唐人蓝色'},
            {'product_keyword': '唐人', 'spec_color_keyword': '粉色', 'target_flower': '唐人粉色'},
            {'product_keyword': '唐人', 'spec_color_keyword': '黄色', 'target_flower': '黄色小唐人'},
            {'product_keyword': '唐人', 'spec_color_keyword': '皮粉', 'target_flower': '唐人粉色'},
            {'product_keyword': '黄底哪吒', 'target_flower': '黄底哪吒'},
            {'product_keyword': '条纹哪吒', 'target_flower': '条纹哪吒'},
            {'product_keyword': '哪吒', 'target_flower': 'Q版哪吒'},
            {'product_keyword': '东北大花', 'spec_keyword': '红底红花', 'target_flower': '东北红底红花'},
            {'product_keyword': '东北大花', 'spec_keyword': '绿底红花', 'target_flower': '东北绿底红花'},
            {'product_keyword': '城市条纹', 'target_flower': '城市条纹'},
            {'product_keyword': '黑白城市条纹', 'target_flower': '城市条纹'},
            {'product_keyword': '油墨风银杏', 'target_flower': '油墨风银杏'},
            {'product_keyword': '银杏叶', 'target_flower': '油墨风银杏'},
            {'product_keyword': '雕印', 'target_flower': '雕印绿色海浪纹'},
            {'product_keyword': '七彩数码印花', 'target_flower': '3D数码印花'},
            {'product_keyword': '3D太阳花', 'target_flower': '3D立体太阳花'},
        ]
        for idx in df[unmatched_mask].index:
            product = str(df.loc[idx, 'product'])
            spec = str(df.loc[idx, 'product_spec'])
            matched = False
            for rule in mapping_rules:
                if 'product_keyword' in rule and rule['product_keyword'] not in product:
                    continue
                if 'spec_color_keyword' in rule and rule['spec_color_keyword'] not in spec:
                    continue
                if 'spec_keyword' in rule and rule['spec_keyword'] not in spec:
                    continue
                target = rule['target_flower']
                if target in cost_flowers:
                    df.loc[idx, '花型'] = target
                    matched = True
                    break
            if not matched and '七彩数码印花' in product:
                if '雕印' in product or '绿色海浪' in product:
                    df.loc[idx, '花型'] = '雕印绿色海浪纹'
                else:
                    df.loc[idx, '花型'] = '3D数码印花'
                matched = True

    df['花型'] = df['花型'].fillna('未匹配')
    return df


def get_order_key(row):
    """同一单（合并发货）的唯一标识：抖音按主订单编号，拼多多/淘宝按快递单号。
    单号为空/NaN 时回落为 order_no（每单单独计费）。"""
    if row.get('platform') == PLATFORM_DOUYIN:
        key = row.get('parent_order_no')
    else:
        key = row.get('express_no')
    if key is None or pd.isna(key) or (isinstance(key, str) and not key.strip()):
        return row.get('order_no')
    return key


def _match_flowers_and_calc(df, cost_map):
    """花型四层匹配 + 米数/成本/快递费/盈利 + 退款标记 + 平台名。

    返回 (df, matched_count, unmatched_count)；无匹配订单时 df 为 None。
    """
    cost_flowers = set(cost_map.keys())
    df = assign_flowers(df, cost_flowers)

    # 打印未匹配订单
    final_unmatched = df[df['花型'] == '未匹配']
    if not final_unmatched.empty:
        print("\n" + "=" * 80)
        print("🚫 所有匹配策略均未成功的订单明细（已从日报中过滤，请人工核查）：")
        print("-" * 80)
        print(final_unmatched[['order_no', 'product', 'product_spec']].to_string(index=False))
        print("=" * 80 + "\n")
    else:
        print("✅ 所有订单均成功匹配花型！")

    matched_df = df[df['花型'].isin(cost_flowers)].copy()
    unmatched_df = df[~df['花型'].isin(cost_flowers)].copy()
    matched_count = len(matched_df)
    unmatched_count = len(unmatched_df)
    print(f"✅ 匹配成功: {matched_count} 条")
    if unmatched_count > 0:
        print(f"⚠️ 未匹配（将被过滤）: {unmatched_count} 条")

    if matched_df.empty:
        return None, matched_count, unmatched_count

    df = matched_df

    # 平台显示名（0=拼多多, 1=淘宝, 2=抖音；旧数据默认拼多多）
    df['平台'] = df['platform'].map(PLATFORM_NAMES).fillna('拼多多')

    # ============================================================
    # 米数计算
    # ============================================================
    df['单位米数'] = df['product_spec'].apply(extract_meter_from_spec)
    df.loc[df['单位米数'] == 0, '单位米数'] = 1
    df['米数'] = (df['单位米数'] * df['product_quantity']).round(2)
    # 保留0.5米倍数
    df['米数'] = (
            (df['米数'] * 2)
            .round()
            / 2
    ).round(2)
    # 检查非0.5倍数米数
    check = df[
        (df['米数'] * 2 % 1 != 0)
    ]

    if not check.empty:
        print("\n==========异常米数订单==========")
        print(
            check[
                [
                    'product_spec',
                    '单位米数',
                    'product_quantity',
                    '米数'
                ]
            ].to_string(index=False)
        )

    # ============================================================
    # 成本计算
    # ============================================================
    df['单位成本'] = df['花型'].apply(lambda x: cost_map.get(x, 0))
    df['成本'] = (df['单位成本'] * df['米数']).round(2)

    # ============================================================
    # 快递费和盈利
    # ============================================================
    # 发货包号：同一 order_key 属同一包裹（合并发货），按首次出现顺序编号
    df['order_key'] = df.apply(get_order_key, axis=1)
    seen, key_to_no, pkg_no = set(), {}, 0
    for key in df['order_key']:
        if key not in seen:
            seen.add(key)
            pkg_no += 1
            key_to_no[key] = pkg_no
    df['发货包号'] = df['order_key'].map(key_to_no)

    # 快递费按包计费：只记在该包第一行，其余行 0，保证汇总不重复计算
    df['快递费'] = 0.0
    for key, group in df.groupby('order_key', sort=False, dropna=False):
        first_idx = group.index[0]
        first_postage = group.iloc[0].get('postage')
        fee = POSTAGE_PER_ORDER
        try:
            if first_postage is not None and float(first_postage) > 0:
                fee = float(first_postage)
        except (TypeError, ValueError):
            pass
        df.loc[first_idx, '快递费'] = fee
    df = df.drop(columns=['order_key'])
    df['盈利'] = (df['merchant_income'] - df['成本'] - df['快递费']).round(2)

    # ============================================================
    # 退款标记
    # ============================================================
    df['是否退款'] = df['after_sale_status'].astype(str).str.contains('退款成功', na=False)

    return df, matched_count, unmatched_count


# ============================
# 汇总表构建（日报/区间报告/月报共用）
# ============================
def _wide_columns(platform_names):
    """左右分列宽表列名：花型 + 每平台6指标(间空2列) + 汇总6指标。"""
    cols = ['花型']
    for i, p in enumerate(platform_names):
        cols += [f'{p}_{m}' for m in SUMMARY_METRICS]
        cols += [f'_spacer_{i}_1', f'_spacer_{i}_2']
    cols += [f'汇总_{m}' for m in SUMMARY_METRICS]
    return cols


def build_platform_summary(normal_df, platform_names=SUMMARY_PLATFORMS, total_label='【总计】'):
    """构建左右分列的宽表汇总。

    行 = 花型（按汇总营业额降序）+ 末尾总计行；
    列 = 花型 | 拼多多6列 | 空2列 | 淘宝6列 | 空2列 | 汇总6列。
    返回扁平列名的 DataFrame，合并表头由 write_summary_sheet 写入。
    """
    def block(flowers_df):
        if flowers_df is None or flowers_df.empty:
            return {}
        g = flowers_df.groupby('花型', dropna=False).agg(
            订单数=('order_no', 'nunique'),
            成本=('成本', 'sum'),
            米数=('米数', 'sum'),
            营业额=('merchant_income', 'sum'),
            快递费=('快递费', 'sum'),
            盈利=('盈利', 'sum')
        ).round(2)
        return g.to_dict(orient='index')

    cols = _wide_columns(platform_names)
    if normal_df is None or normal_df.empty:
        return pd.DataFrame(columns=cols)

    total = block(normal_df)
    # 花型按汇总营业额降序
    flower_order = sorted(total.keys(), key=lambda f: total[f]['营业额'], reverse=True)

    per_platform = {p: block(normal_df[normal_df['平台'] == p]) for p in platform_names}

    rows = []
    for flower in flower_order:
        row = {'花型': flower}
        for i, p in enumerate(platform_names):
            fb = per_platform[p].get(flower, {})
            for m in SUMMARY_METRICS:
                row[f'{p}_{m}'] = fb.get(m, 0)
            row[f'_spacer_{i}_1'] = ''
            row[f'_spacer_{i}_2'] = ''
        for m in SUMMARY_METRICS:
            row[f'汇总_{m}'] = total[flower][m]
        rows.append(row)

    # 总计行
    def raw_col(m):
        return 'merchant_income' if m == '营业额' else m

    tot = {'花型': total_label}
    for i, p in enumerate(platform_names):
        sub = normal_df[normal_df['平台'] == p]
        tot[f'{p}_订单数'] = sub['order_no'].nunique() if not sub.empty else 0
        for m in SUMMARY_METRICS[1:]:
            tot[f'{p}_{m}'] = round(sub[raw_col(m)].sum(), 2) if not sub.empty else 0
        tot[f'_spacer_{i}_1'] = ''
        tot[f'_spacer_{i}_2'] = ''
    tot['汇总_订单数'] = len(normal_df)
    for m in SUMMARY_METRICS[1:]:
        tot[f'汇总_{m}'] = round(normal_df[raw_col(m)].sum(), 2)
    rows.append(tot)

    return pd.DataFrame(rows, columns=cols)


def build_platform_totals(normal_df, platform_names=PLATFORM_SUMMARY_NAMES):
    """平台汇总表：指定平台三行齐全（无数据补 0），末尾合计。"""
    def row(name, sub):
        return {
            '平台': name,
            '订单数': sub['order_no'].nunique() if not sub.empty else 0,
            '成本': round(sub['成本'].sum(), 2) if not sub.empty else 0,
            '米数': round(sub['米数'].sum(), 2) if not sub.empty else 0,
            '营业额': round(sub['merchant_income'].sum(), 2) if not sub.empty else 0,
            '快递费': round(sub['快递费'].sum(), 2) if not sub.empty else 0,
            '盈利': round(sub['盈利'].sum(), 2) if not sub.empty else 0,
        }

    if normal_df is None or normal_df.empty:
        rows = [row(n, pd.DataFrame()) for n in platform_names]
        rows.append({'平台': '合计', '订单数': 0, '成本': 0, '米数': 0, '营业额': 0, '快递费': 0, '盈利': 0})
        return pd.DataFrame(rows)

    rows = [row(n, normal_df[normal_df['平台'] == n]) for n in platform_names]
    rows.append({
        '平台': '合计',
        '订单数': len(normal_df),
        '成本': round(normal_df['成本'].sum(), 2),
        '米数': round(normal_df['米数'].sum(), 2),
        '营业额': round(normal_df['merchant_income'].sum(), 2),
        '快递费': round(normal_df['快递费'].sum(), 2),
        '盈利': round(normal_df['盈利'].sum(), 2),
    })
    return pd.DataFrame(rows)


def write_summary_sheet(writer, summary_wide, sheet_name='花型汇总'):
    """把左右分列宽表汇总写入 Excel：第1行平台合并表头，第2行指标名，数据从第3行起，总计行加粗。"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    summary_wide.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=2)
    if sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
    else:
        ws = writer.book.create_sheet(sheet_name)
        writer.sheets[sheet_name] = ws

    cols = list(summary_wide.columns)

    # 按列名前缀分组：'拼多多_订单数' → 平台段 '拼多多'；'_spacer_0_1' → 空段
    def group_key(c):
        if c.startswith('_spacer_'):
            return c.rsplit('_', 1)[0]  # '_spacer_0' 每平台独立空段
        return c.split('_')[0]

    groups = []
    for c in cols[1:]:
        key = group_key(c)
        if groups and groups[-1][0] == key:
            groups[-1][1].append(c)
        else:
            groups.append((key, [c]))

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    title_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')

    # 花型表头（合并 A1:A2）
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1, value='花型')
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = border

    col_idx = 1  # A 列（花型）
    for key, gc in groups:
        start = col_idx + 1
        end = col_idx + len(gc)
        is_platform = not key.startswith('_spacer_')
        # 第1行：平台名（合并段），空段留白
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        if is_platform:
            ws.cell(row=1, column=start, value=key)
        for j in range(start, end + 1):
            cell = ws.cell(row=1, column=j)
            if is_platform:
                cell.font = title_font
                cell.alignment = center
                cell.fill = header_fill
            cell.border = border
        # 第2行：指标名
        for j, c in enumerate(gc):
            cell = ws.cell(row=2, column=start + j)
            if is_platform:
                cell.value = c.split('_', 1)[1]
                cell.font = title_font
                cell.alignment = center
                cell.fill = header_fill
            cell.border = border
        col_idx = end

    # 数据行边框 + 总计行加粗（总计行由 build_platform_summary 追加在最后）
    n_data = len(summary_wide)
    last_data_row = 2 + n_data
    for r in range(3, 3 + n_data):
        is_total = (r == last_data_row)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if is_total:
                cell.font = Font(bold=True)


def _auto_width_sheets(writer):
    from openpyxl.utils import get_column_letter
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        for col_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                if cell.value is None:
                    continue
                try:
                    length = len(str(cell.value))
                except Exception:
                    continue
                if length > max_length:
                    max_length = length
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width


# ============================
# 主函数
# ============================
def generate_daily_report(target_date=None, force=True, orders=orders):
    """
    生成日报（默认强制覆盖）
    force=True：强制重新生成，先回退旧库存，再重新扣减
    """
    ensure_output_dir()

    if target_date is None:
        target_dt = datetime.now()
    else:
        try:
            target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        except ValueError:
            print(f"日期格式错误: {target_date}，请使用 YYYY-MM-DD 格式")
            return

    target_date = target_dt.strftime('%Y-%m-%d')
    filename = f"{target_dt.strftime('%Y%m%d')}日报.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    # ============================================================
    # 🔧 改进：读取旧缓存用于对比，但不回退库存（改用增量扣减）
    # ============================================================
    old_data = None
    with engine.connect() as conn:
        # 1. 检查是否有旧缓存，并读取旧数据用于对比
        existing = conn.execute(
            text("SELECT COUNT(*) FROM daily_report_cache WHERE report_date = :d"),
            {"d": target_date}
        ).scalar()

        if existing > 0:
            old_data = pd.read_sql(
                text("SELECT flower, total_meters, revenue, profit FROM daily_report_cache WHERE report_date = :d"),
                conn,
                params={"d": target_date}
            )
            if not old_data.empty:
                old_data = old_data.set_index('flower').to_dict(orient='index')
            else:
                old_data = {}
            print(f"📊 检测到 {target_date} 已有日报缓存")
        else:
            old_data = {}
            print(f"📝 {target_date} 首次生成日报")

        # 2. 🔧 不再自动回退库存，改为增量扣减（见下方扣减逻辑）
        # 只清理缓存数据，库存变动记录保留用于增量计算

        # 3. 删除所有旧缓存（确保插入时不会唯一键冲突）
        conn.execute(
            text("DELETE FROM daily_report_cache WHERE report_date = :d"),
            {"d": target_date}
        )
        conn.execute(
            text("DELETE FROM daily_report_meta WHERE report_date = :d"),
            {"d": target_date}
        )
        conn.execute(
            text("DELETE FROM inventory_shortfall WHERE reference_date = :d"),
            {"d": target_date}
        )
        conn.commit()
        print(f"✅ 已清理 {target_date} 的所有旧缓存数据")

    # ============================================================
    # 读取当天订单
    # ============================================================
    query = text(f"""
        SELECT
            id, order_no, product, product_spec,
            product_quantity, merchant_income,
            cost, meter, express_cost, traffic_cost, profit,
            after_sale_status, order_status, platform,
            express_no, postage, parent_order_no
        FROM {orders}
        WHERE DATE(delivery_time) = :target_date
          AND delivery_time IS NOT NULL
    """)

    try:
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params={"target_date": target_date})
    except Exception as e:
        print(f"读取数据库失败: {e}")
        return

    if df.empty:
        print(f"⚠️ {target_date} 没有订单数据，无需生成日报")
        return

    original_count = len(df)
    print(f"📋 原始订单数: {original_count}")

    # 加载成本表
    cost_map = load_cost_map(target_date)
    cost_flowers = set(cost_map.keys())
    print(f"📚 成本表中有 {len(cost_map)} 个花型")

    # ============================================================
    # 花型匹配 + 米数/成本/快递费/盈利 + 退款标记（共用函数，日报与区间报告口径一致）
    # ============================================================
    df, matched_count, unmatched_count = _match_flowers_and_calc(df, cost_map)
    if df is None:
        print(f"❌ 没有匹配成功的订单，无法生成日报")
        return

    # ============================================================
    # 正常订单（汇总表用：排除所有退款和取消）
    # ============================================================
    normal_df = df[(~df['是否退款']) & (~df['order_status'].isin(CANCELLED_STATUSES))]

    # ============================================================
    # 明细表数据：正常订单 + 已发货/已收货退款订单（用于对账/运费核算）
    # ============================================================
    # 需要保留的退款状态（涉及运费承担）
    keep_refund_statuses = [
        '已发货，退款成功',
        '已收货，退款成功'
    ]

    # 明细表数据：正常订单 + 符合条件的退款订单
    detail_df = df[
        (~df['order_status'].isin(CANCELLED_STATUSES)) &
        (
                (~df['是否退款']) |
                (df['order_status'].isin(keep_refund_statuses))
        )
        ].copy()

    # 退款订单的成本和米数设为 0
    detail_df.loc[detail_df['是否退款'] == True, ['成本', '米数']] = 0

    # ============================================================
    # 明细表
    # ============================================================
    detail_cols = ['花型', '平台', '发货包号', '成本', '米数', 'merchant_income', '快递费', '盈利',
                   'after_sale_status', 'order_no', 'product_spec', 'product_quantity', '是否退款']
    detail = detail_df[detail_cols].copy()
    detail = detail.rename(columns={
        'merchant_income': '营业额',
        'after_sale_status': '售后状态',
        'product_spec': '商品规格',
        'product_quantity': '商品数量'
    })
    detail = detail.sort_values('花型')

    # ============================================================
    # 写入日报缓存
    # ============================================================
    if not normal_df.empty:
        with engine.connect() as conn:
            cache_data = normal_df.groupby('花型').agg(
                order_count=('order_no', 'nunique'),
                total_meters=('米数', 'sum'),
                revenue=('merchant_income', 'sum'),
                cost=('成本', 'sum'),
                express_fee=('快递费', 'sum'),
                profit=('盈利', 'sum')
            ).reset_index()

            for _, row in cache_data.iterrows():
                conn.execute(
                    text("""
                        INSERT INTO daily_report_cache 
                        (report_date, flower, order_count, total_meters, revenue, cost, express_fee, profit)
                        VALUES (:d, :f, :oc, :tm, :rev, :cst, :exp, :prf)
                    """),
                    {
                        "d": target_date,
                        "f": row['花型'],
                        "oc": int(row['order_count']),
                        "tm": float(row['total_meters']),
                        "rev": float(row['revenue']),
                        "cst": float(row['cost']),
                        "exp": float(row['express_fee']),
                        "prf": float(row['profit'])
                    }
                )
            conn.commit()
            print(f"✅ 日报缓存已写入：{len(cache_data)} 个花型")

    # ============================================================
    # 变化报告（对比新旧数据）
    # ============================================================
    if old_data and not normal_df.empty:
        print("\n" + "=" * 60)
        print(f"📊 {target_date} 日报变化报告：")
        print("-" * 60)

        new_data = normal_df.groupby('花型').agg(
            total_meters=('米数', 'sum'),
            revenue=('merchant_income', 'sum'),
            profit=('盈利', 'sum')
        ).to_dict(orient='index')

        all_flowers = set(old_data.keys()) | set(new_data.keys())
        changes = []
        for flower in sorted(all_flowers):
            old = old_data.get(flower, {})
            new = new_data.get(flower, {})
            old_m = old.get('total_meters', 0)
            new_m = new.get('total_meters', 0)
            old_r = old.get('revenue', 0)
            new_r = new.get('revenue', 0)
            old_p = old.get('profit', 0)
            new_p = new.get('profit', 0)

            if abs(old_m - new_m) > 0.01 or abs(old_r - new_r) > 0.01:
                changes.append({
                    '花型': flower,
                    '旧米数': old_m,
                    '新米数': new_m,
                    '米数变化': new_m - old_m,
                    '旧营业额': old_r,
                    '新营业额': new_r,
                    '营业额变化': new_r - old_r,
                    '旧利润': old_p,
                    '新利润': new_p,
                    '利润变化': new_p - old_p
                })

        if changes:
            change_df = pd.DataFrame(changes)
            print(change_df.to_string(index=False))
            print(f"\n📌 共有 {len(changes)} 个花型发生变化")
        else:
            print("✅ 数据无变化（但已强制重新生成）")

        print("=" * 60 + "\n")
    elif old_data and normal_df.empty:
        print(f"⚠️ {target_date} 当前无正常订单，旧数据已被清除")

    # ============================================================
    # 🔧 回退旧扣减 + 重新扣减库存（不依赖 inventory_log）
    # ============================================================
    # 第 1 步：根据旧缓存数据回退上一次日报的扣减
    if old_data:
        rollback_count = 0
        with engine.connect() as conn:
            trans = conn.begin()
            for flower, data in old_data.items():
                prev_meters = float(data.get('total_meters', 0))
                if prev_meters <= 0.001:
                    continue
                # 读回退前快照值
                before_row = conn.execute(
                    text("SELECT stock FROM inventory_snapshot WHERE flower = :f AND snapshot_date = :d"),
                    {"f": flower, "d": target_date}
                ).fetchone()
                before_stock = float(before_row[0]) if before_row else 0

                # 回退实时库存
                conn.execute(
                    text("UPDATE inventory SET current_stock = current_stock + :qty WHERE flower = :f"),
                    {"qty": prev_meters, "f": flower}
                )
                # 回退快照（从 target_date 起所有快照 +prev_meters）
                conn.execute(
                    text("""
                        UPDATE inventory_snapshot
                        SET stock = stock + :qty, updated_by = 'system', updated_at = CURRENT_TIMESTAMP
                        WHERE flower = :f AND snapshot_date >= :d
                    """),
                    {"qty": prev_meters, "f": flower, "d": target_date}
                )
                # 写入回退日志
                conn.execute(
                    text("""
                        INSERT INTO inventory_log
                        (flower, change_type, change_qty, before_stock, after_stock, reference, operator)
                        VALUES (:f, '手动调整', :qty, :before, :after, :ref, :op)
                    """),
                    {
                        "f": flower, "qty": prev_meters,
                        "before": before_stock,
                        "after": before_stock + prev_meters,
                        "ref": f"回退日报 {target_date}（重新生成）",
                        "op": "system"
                    }
                )
                rollback_count += 1
            trans.commit()
        if rollback_count > 0:
            print(f"🔄 已回退 {rollback_count} 个花型的旧扣减")

    # 第 2 步：按新日报重新扣减
    if not normal_df.empty:
        sales_summary = normal_df.groupby('花型')['米数'].sum().reset_index()
        print("\n📦 正在扣减库存...")
        deducted_count = 0
        for _, row in sales_summary.iterrows():
            flower = row['花型']
            current_meters = float(row['米数'])
            if current_meters <= 0.001:
                continue
            try:
                deduct_stock(
                    flower=flower,
                    qty=current_meters,
                    reference=f"日报自动扣减 {target_date}",
                    operator="system",
                    report_date=target_date
                )
                deducted_count += 1
                print(f"  ✅ {flower}: 扣减 {current_meters:.1f} 米")
            except ValueError as e:
                print(f"  ⚠️ {e}，跳过该花型")

        if deducted_count > 0:
            print(f"✅ 库存扣减完成：{deducted_count} 个花型")
        else:
            print("ℹ️ 没有需要扣减的花型")
    elif not old_data:
        print("ℹ️ 无旧数据且无新订单，无需库存变动")

    # ============================================================
    # 查询缺口
    # ============================================================
    with engine.connect() as conn:
        gaps = conn.execute(
            text("""
                SELECT flower, shortfall_meters 
                FROM inventory_shortfall 
                WHERE reference_date = :d AND status = '待补录'
            """),
            {"d": target_date}
        ).fetchall()

    if gaps:
        print("\n" + "=" * 50)
        print("📋 以下花型库存不足，已产生缺口（请尽快补录）：")
        for row in gaps:
            print(f"   {row[0]}: 缺口 {row[1]} 米")
        print("=" * 50)
    else:
        print("✅ 当天所有花型库存充足，无缺口")

    # ============================================================
    # 汇总表（左右分列：花型 | 拼多多6列 | 空2列 | 淘宝6列 | 空2列 | 汇总6列）
    # ============================================================
    summary_wide = build_platform_summary(normal_df)
    platform_summary = build_platform_totals(normal_df)




    # 打印总计对比
    detail_total_cost = detail['成本'].sum()
    detail_total_meter = detail['米数'].sum()
    summary_total_cost = normal_df['成本'].sum() if not normal_df.empty else 0
    summary_total_meter = normal_df['米数'].sum() if not normal_df.empty else 0
    print(f"\n总计对比：")
    print(f"  明细总成本: {detail_total_cost:.2f}  汇总总成本: {summary_total_cost:.2f}  差异: {summary_total_cost - detail_total_cost:.2f}")
    print(f"  明细总米数: {detail_total_meter:.2f}  汇总总米数: {summary_total_meter:.2f}  差异: {summary_total_meter - detail_total_meter:.2f}")
    print("=" * 80 + "\n")





    # ============================================================
    # 保存 Excel
    # ============================================================
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        write_summary_sheet(writer, summary_wide, sheet_name='花型汇总')
        platform_summary.to_excel(writer, sheet_name='平台汇总', index=False)
        detail.to_excel(writer, sheet_name='订单明细', index=False)
        _auto_width_sheets(writer)

    print(f"✅ 日报已生成: {filepath}")
    if not normal_df.empty:
        print(f"📊 正常订单 {len(normal_df)} 单，总营业额 {normal_df['merchant_income'].sum():.2f}，总盈利 {normal_df['盈利'].sum():.2f}")
    if unmatched_count > 0:
        print(f"⚠️ 已过滤 {unmatched_count} 条未匹配花型的订单")
    # ============================================================
    # 🆕 生成库存快照
    # ============================================================
    print("\n📸 正在生成库存快照...")
    from inventory_service import fill_missing_snapshots, get_missing_report_dates

    # 补全缺失的快照（从上次快照到当天）
    filled, status, msg = fill_missing_snapshots(up_to_date=None, operator="system")

    if status == "no_report":
        print(f"⚠️ {msg}")
    elif status == "already_latest":
        print(f"ℹ️ {msg}")
    elif status == "error":
        print(f"❌ {msg}")
    else:
        print(f"✅ {msg}")

    # 检查是否有缺失的日报（自动提醒）
    missing_reports = get_missing_report_dates()
    if missing_reports:
        print(f"⚠️ 以下日期有订单但未生成日报：{missing_reports}")
    return filepath


# ============================
# 区间报告函数
# ============================
def generate_range_report(start_date, end_date, force=True, orders=orders):
    """生成指定日期区间的汇总报表（只输出 Excel，不触碰库存/日报缓存）。

    数据口径与日报一致：只要发货/收货且不退款，都计入营业额与利润；
    仅排除退款订单与已取消/已关闭订单。
    文件名：区间报告_YYYYMMDD-YYYYMMDD.xlsx
    """
    ensure_output_dir()

    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    except ValueError:
        print(f"❌ 日期格式错误: {start_date} / {end_date}，请使用 YYYY-MM-DD 格式")
        return None

    if start_dt > end_dt:
        print("❌ 开始日期不能晚于结束日期")
        return None

    start_date = start_dt.strftime('%Y-%m-%d')
    end_date = end_dt.strftime('%Y-%m-%d')
    filename = f"区间报告_{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    if os.path.exists(filepath) and not force:
        print(f"⏭️ {filename} 已存在，跳过生成")
        return filepath

    # ============================================================
    # 读取区间订单
    # ============================================================
    query = text(f"""
        SELECT
            id, order_no, product, product_spec,
            product_quantity, merchant_income,
            cost, meter, express_cost, traffic_cost, profit,
            after_sale_status, order_status, platform,
            express_no, postage, parent_order_no
        FROM {orders}
        WHERE delivery_time >= :start_date
          AND delivery_time < DATE_ADD(:end_date, INTERVAL 1 DAY)
          AND delivery_time IS NOT NULL
    """)
    try:
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params={"start_date": start_date, "end_date": end_date})
    except Exception as e:
        print(f"读取数据库失败: {e}")
        return None

    if df.empty:
        print(f"⚠️ {start_date} 到 {end_date} 没有订单数据")
        return None

    print(f"📋 区间订单数: {len(df)} 条")

    # 加载成本表（按区间结束日口径）
    cost_map = load_cost_map(end_date)
    print(f"📚 成本表中有 {len(cost_map)} 个花型（按 {end_date} 口径）")

    # 花型匹配 + 计算（与日报共用）
    df, matched_count, unmatched_count = _match_flowers_and_calc(df, cost_map)
    if df is None:
        print("❌ 区间内没有匹配成功的订单")
        return None

    # ============================================================
    # 汇总口径（与日报一致）
    # ============================================================
    normal_df = df[(~df['是否退款']) & (~df['order_status'].isin(CANCELLED_STATUSES))]

    keep_refund_statuses = [
        '已发货，退款成功',
        '已收货，退款成功'
    ]
    detail_df = df[
        (~df['order_status'].isin(CANCELLED_STATUSES)) &
        (
                (~df['是否退款']) |
                (df['order_status'].isin(keep_refund_statuses))
        )
        ].copy()
    detail_df.loc[detail_df['是否退款'] == True, ['成本', '米数']] = 0

    # 明细表
    detail_cols = ['花型', '平台', '发货包号', '成本', '米数', 'merchant_income', '快递费', '盈利',
                   'after_sale_status', 'order_no', 'product_spec', 'product_quantity', '是否退款']
    detail = detail_df[detail_cols].copy()
    detail = detail.rename(columns={
        'merchant_income': '营业额',
        'after_sale_status': '售后状态',
        'product_spec': '商品规格',
        'product_quantity': '商品数量'
    })
    detail = detail.sort_values('花型')

    # 汇总表
    summary_wide = build_platform_summary(normal_df)
    platform_summary = build_platform_totals(normal_df)

    if not normal_df.empty:
        print(f"📊 正常订单 {len(normal_df)} 单 / 营业额 {normal_df['merchant_income'].sum():.2f} / 盈利 {normal_df['盈利'].sum():.2f}")
    else:
        print("⚠️ 区间内没有正常订单（可能全部退款/取消）")

    # ============================================================
    # 保存 Excel
    # ============================================================
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        write_summary_sheet(writer, summary_wide, sheet_name='花型汇总')
        platform_summary.to_excel(writer, sheet_name='平台汇总', index=False)
        detail.to_excel(writer, sheet_name='订单明细', index=False)
        _auto_width_sheets(writer)

    print(f"✅ 区间报告已生成: {filepath}")
    return filepath


# ============================
# 批量补全函数
# ============================
def generate_all_missing_reports(orders=orders):
    ensure_output_dir()
    # 1. 获取所有有订单的日期（按 delivery_time）
    query_order_dates = text(f"""
        SELECT DISTINCT DATE(delivery_time) AS order_date
        FROM {orders}
        WHERE delivery_time IS NOT NULL
        ORDER BY order_date
    """)
    # 2. 获取已有日报缓存的日期
    query_report_dates = text("""
        SELECT DISTINCT report_date FROM daily_report_cache
    """)

    with engine.connect() as conn:
        df_order_dates = pd.read_sql(query_order_dates, conn)
        df_report_dates = pd.read_sql(query_report_dates, conn)

    if df_order_dates.empty:
        print("没有订单数据")
        return

    order_dates = df_order_dates['order_date'].dropna().tolist()
    report_dates = df_report_dates['report_date'].dropna().tolist() if not df_report_dates.empty else []

    # 3. 只生成「有订单但无日报缓存」的日期，已有日报的跳过，避免重复生成
    missing_dates = sorted(set(order_dates) - set(report_dates))

    if not missing_dates:
        print("所有有订单的日期均已有日报，无需生成")
        return

    print(f"找到 {len(missing_dates)} 个缺失日报的日期")
    generated = 0
    failed = []
    for date in missing_dates:
        date_ymd = date.strftime('%Y-%m-%d')
        print(f"  📊 生成 {date_ymd} 日报...")
        try:
            generate_daily_report(date_ymd, force=True)
            generated += 1
        except OSError as e:
            # Windows 下目标日报文件可能正被 Excel 打开占用，跳过该日期继续生成其余日期
            failed.append(date_ymd)
            print(f"  ⚠️ {date_ymd} 日报写入失败（文件被占用/锁定）: {e}")
            print(f"     请关闭 {OUTPUT_DIR}\\{date_ymd.replace('-', '')}日报.xlsx 后重试该日期")
    print(f"\n✅ 完成: 成功生成 {generated} 个")
    if failed:
        print(f"⚠️ 以下日期写入失败（通常因文件正被 Excel 打开）: {', '.join(failed)}")

# ============================
# 入口
# ============================
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        if sys.argv[1] == "--all":
            generate_all_missing_reports()
        elif sys.argv[1] == "--range" and len(sys.argv) == 4:
            start = sys.argv[2]
            end = sys.argv[3]
            current = datetime.strptime(start, '%Y-%m-%d')
            end_dt = datetime.strptime(end, '%Y-%m-%d')
            while current <= end_dt:
                generate_daily_report(current.strftime('%Y-%m-%d'), force=True)
                current += timedelta(days=1)
        elif sys.argv[1] == "--report" and len(sys.argv) == 4:
            generate_range_report(sys.argv[2], sys.argv[3], force=True)
        else:
            generate_daily_report(sys.argv[1], force=True)
    else:
            # 默认生成指定日期的日报（可自行修改日期）
            generate_daily_report("2026-07-01")