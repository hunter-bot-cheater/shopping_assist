"""Tests for make_report.generate_inout_detail_report.

The engine is mocked at module level (conftest.py); this test drives
engine.connect through a per-query side_effect and inspects the generated
Workbook written to a tmp directory.
"""
import os
from datetime import date
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import make_report
from make_report import generate_inout_detail_report


def make_side_effect(flowers, init_rows, out_rows, in_rows):
    def side_effect(sql, params=None, **kw):
        s = str(sql)
        res = MagicMock()
        if 'product_cost' in s and 'is_deleted = 0' in s:
            res.fetchall.return_value = flowers
        elif 'inventory_snapshot' in s:
            res.fetchall.return_value = init_rows
        elif 'daily_report_cache' in s:
            res.fetchall.return_value = out_rows
        elif 'inventory_log' in s:
            res.fetchall.return_value = in_rows
        else:
            res.fetchall.return_value = []
        return res
    return side_effect


@pytest.fixture
def mock_engine_conn():
    with patch('make_report.engine.connect') as m:
        ctx = MagicMock()
        conn = MagicMock()
        ctx.__enter__.return_value = conn
        m.return_value = ctx
        yield conn


def test_generate_report_structure(tmp_path, mock_engine_conn):
    """每个花型占 2 行，期初/花型/余额合并，出库总数/入库总数正确，格式符合样例。"""
    mock_engine_conn.execute.side_effect = make_side_effect(
        flowers=[('花型A',), ('花型B',)],
        init_rows=[('花型A', 100.0)],
        out_rows=[
            (date(2026, 7, 1), '花型A', 10.0),
            (date(2026, 7, 2), '花型A', 5.0),
            (date(2026, 7, 1), '花型B', 3.0),
        ],
        in_rows=[(date(2026, 7, 1), '花型A', 20.0)],
    )
    with patch.object(make_report, 'OUTPUT_DIR', str(tmp_path)):
        fp = generate_inout_detail_report('2026-07-01', '2026-07-02')

    assert os.path.basename(fp) == '260701-260702出库入库明细.xlsx'

    ws = load_workbook(fp)['出库入库明细']
    ncols = ws.max_column
    out_col, bal_col = ncols - 1, ncols
    out_letter, bal_letter = get_column_letter(out_col), get_column_letter(bal_col)

    # 标题合并整行、表头
    assert ws['A1'].value == '260701-260702出库入库明细'
    assert ws['B2'].value == '日期'
    assert ws['D2'].value == 1
    assert ws['E2'].value == 2
    assert ws['A3'].value == '6月30日库存'
    assert ws['B3'].value == '花型名称'
    assert ws[f'{out_letter}3'].value == '总出入库'
    assert ws[f'{bal_letter}3'].value == '库存余额'

    # 花型A：期初100，7/1出库10 7/2出库5，出库总数15，入库7/1=20，余额=100+20-15=105
    assert ws['A4'].value == 100.0
    assert ws['B4'].value == '花型A'
    assert ws['C4'].value == '出库米数'
    assert ws['C5'].value == '入库米数'
    assert ws['D4'].value == 10.0
    assert ws['E4'].value == 5.0
    assert ws[f'{out_letter}4'].value == 15.0
    assert ws['D5'].value == 20.0
    assert ws[f'{out_letter}5'].value == 20.0
    assert ws[f'{bal_letter}4'].value == 105.0
    # 合并与格式
    merges = {str(r) for r in ws.merged_cells.ranges}
    assert 'A4:A5' in merges and 'B4:B5' in merges and f'{bal_letter}4:{bal_letter}5' in merges
    assert ws['D4'].number_format == '0.0'
    assert ws[f'{out_letter}4'].font.color.rgb == '00FF0000'
    assert ws['A3'].fill.fgColor.rgb == '00D9E1F2'

    # 固定紧凑列宽
    assert ws.column_dimensions['A'].width == 11
    assert ws.column_dimensions['B'].width == 12
    assert ws.column_dimensions['D'].width == 6
    assert ws.column_dimensions[out_letter].width == 8
    assert ws.column_dimensions[bal_letter].width == 10

    # 花型B：无期初(0)，7/1出库3，余额=0-3=-3
    assert ws['A6'].value == 0.0
    assert ws['B6'].value == '花型B'
    assert ws[f'{out_letter}6'].value == 3.0
    assert ws[f'{bal_letter}6'].value == -3.0


def test_start_after_end_raises(mock_engine_conn):
    mock_engine_conn.execute.side_effect = make_side_effect([('A',)], [], [], [])
    with pytest.raises(ValueError, match='开始日期不能晚于结束日期'):
        generate_inout_detail_report('2026-07-02', '2026-07-01')


def test_no_movement_raises(tmp_path, mock_engine_conn):
    """范围内无出库无入库 → 提示无数据。"""
    mock_engine_conn.execute.side_effect = make_side_effect(
        flowers=[('花型A',)], init_rows=[], out_rows=[], in_rows=[]
    )
    with patch.object(make_report, 'OUTPUT_DIR', str(tmp_path)):
        with pytest.raises(ValueError, match='所选日期范围内无数据'):
            generate_inout_detail_report('2026-07-01', '2026-07-02')


def test_no_flowers_raises(mock_engine_conn):
    mock_engine_conn.execute.side_effect = make_side_effect([], [], [], [])
    with pytest.raises(ValueError, match='请先添加花型'):
        generate_inout_detail_report('2026-07-01', '2026-07-02')
